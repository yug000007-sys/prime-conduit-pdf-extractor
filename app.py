import re
from io import BytesIO

import pdfplumber
import streamlit as st
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


st.set_page_config(page_title="Prime Conduit PDF Extractor", page_icon="📊", layout="wide")

# Exact Book5-style header.
# Unwanted data columns remain blank:
# meta_data_json, SO_Number, Commission_Rate, currency, rawcustname
HEADERS = ['Distname', 'Supplier_name', 'direct_indirect', 'in_out_territory', 'CustAccNbr', 'CustDunsID', 'CustName', 'Address1', 'City', 'State', 'County', 'Zip', 'Phone', 'Country', 'NoOfEmployees', 'WebAddress', 'SIC', 'NAICS', 'LineOfBusiness', 'ParentName', 'AccountType', 'UOM', 'InvoiceNumber', 'Qty', 'UnitCost', 'UnitResale', 'InvoiceDate', 'DateRecieved', 'PartNumberSubmitted', 'PartNumberDescription', 'Branch', 'SalesRep', 'Latitude', 'Longitude', 'Brand', 'PartNumberActual', 'UPCCode', 'rawcustname', 'rawdistaddress', 'rawdistcity', 'rawdiststate', 'rawdistpostalcode', 'rawdistcountry', 'currency', 'contractID', 'client_CustName', 'Zip_4_digit', 'dnb_trade_style', 'dnb_sales_value', 'google_CustName', 'google_Address1', 'google_State', 'google_Zip', 'google_Country', 'google_Phone', 'google_WebAddress', 'Pay_Month', 'Pay_Year', 'Ship_Month', 'Ship_Year', 'Industry', 'Commissions', 'Commission_Rate', 'Cust_AM', 'CEM', 'Sales', 'In_Out', 'Commission_split_percentage', 'Distributor_part_number', 'Category', 'google_City', 'Billings', 'Cheque_Number', 'Pay_Date', 'meta_data_json', 'SO_Number', 'PO_Number', 'ship_date', 'searched_on_google']

INVOICE_RE = re.compile(r"\b9\d{7,8}\b")
ACCOUNT_RE = re.compile(r"^(\d{5,6})\s*-\s*(.+)$")
STATE_RE = re.compile(
    r"\b(AL|AK|AZ|AR|CA|CO|CT|DE|FL|GA|IA|ID|IL|IN|KS|KY|LA|MA|MD|ME|MI|MN|MO|MS|MT|NC|ND|NE|NH|NJ|NM|NV|NY|OH|OK|OR|PA|RI|SC|SD|TN|TX|UT|VA|VT|WA|WI|WV|WY)\b"
)
AMOUNT_RE = re.compile(r"^-?\d{1,3}(?:,\d{3})*(?:\.\d{2})-?$|^-?\d+(?:\.\d{2})-?$")
DATE_RE = re.compile(
    r"Commission Customer Report For Period\s*:\s*(\d{2})/(\d{2})/(\d{4})\s*to\s*(\d{2})/(\d{2})/(\d{4})"
)

SKIP_WORDS = (
    "ZSDB0010", "Commission Customer Report", "Commission Group Report",
    "Agency Name", "Agency No.", "Agreement No.", "Sales Organization",
    "Distribution Channel", "Customer Name", "City / State", "Ship To Party",
    "Customer Totals", "Commission Totals", "Group Description", "Frt Allow",
    "Freight / Basis", "Page :", "Time :"
)


def clean(value):
    return re.sub(r"\s+", " ", str(value or "")).strip()


def to_float(value, make_positive=True):
    value = clean(value).replace("$", "").replace(",", "")

    if value.endswith("-"):
        value = "-" + value[:-1]

    if not value:
        return ""

    try:
        number = float(value)
        return abs(number) if make_positive else number
    except ValueError:
        return value


def normalize_customer_name(group_name):
    """
    Book5 uses the main customer/distributor block name as CustName,
    not the ship-to description after the account number.

    Examples:
    CED 4662 -> CED
    WESCO - FARGO ND 7852 -> WESCO - FARGO ND
    VIKING ELECTRIC SPLY INC -> VIKING ELECTRIC SPLY INC
    """
    name = clean(group_name)
    name = re.sub(r"\s+\d{3,}$", "", name).strip()
    return name


def parse_report_period(text):
    match = DATE_RE.search(text)
    if not match:
        return "", "", "", ""

    start_month, start_day, start_year, end_month, end_day, end_year = match.groups()
    pay_month = end_month
    pay_year = end_year
    ship_month = end_month
    ship_year = end_year
    return pay_month, pay_year, ship_month, ship_year


def get_amount_tokens(text_after_invoice):
    tokens = clean(text_after_invoice).split()

    # Ignore percentages, keep only money/amount columns.
    return [
        token for token in tokens
        if AMOUNT_RE.match(token) and not token.endswith("%")
    ]


def parse_invoice_line(line):
    match = INVOICE_RE.search(line)
    if not match:
        return None

    invoice_number = match.group(0)
    group_name_on_line = clean(line[:match.start()])
    text_after_invoice = clean(line[match.end():])
    amounts = get_amount_tokens(text_after_invoice)

    net_value = ""
    commission_basis = ""
    commission_total = ""

    # Normal report line:
    # Invoice NetValue Discount/Freight CommissionBasis Rate CommissionTotal
    if len(amounts) >= 5:
        net_value = to_float(amounts[0])
        commission_basis = to_float(amounts[2])
        commission_total = to_float(amounts[-1])

    # Telecom-style line may not have Discount/Freight:
    # Invoice NetValue CommissionBasis Rate CommissionTotal
    elif len(amounts) >= 3:
        net_value = to_float(amounts[0])
        commission_basis = to_float(amounts[-2])
        commission_total = to_float(amounts[-1])

    elif amounts:
        commission_total = to_float(amounts[-1])

    return {
        "invoice_number": invoice_number,
        "group_name_on_line": group_name_on_line,
        "net_value": net_value,
        "commission_basis": commission_basis,
        "commission_total": commission_total,
    }


def parse_city_state_order_line(line):
    """
    Handles both:
    MINNEAPOLIS MN 2368091 P001829508 1,089.84
    2368092 P001829516 1,910.06

    Book5 maps PO_Number to the order number, not the actual P.O. number.
    SO_Number stays blank.
    """
    line = clean(line)
    city = ""
    state = ""
    order_number = ""

    state_match = STATE_RE.search(line)
    remaining = line

    if state_match:
        city = clean(line[:state_match.start()])
        state = state_match.group(1)
        remaining = clean(line[state_match.end():])

    parts = remaining.split()

    if parts and re.fullmatch(r"\d{6,8}", parts[0]):
        order_number = parts[0]

    return city, state, order_number


def find_ship_to_account(lines, start_index):
    for index in range(start_index, min(start_index + 8, len(lines))):
        match = ACCOUNT_RE.search(clean(lines[index]))
        if match:
            return match.group(1)
    return ""


def extract_rows_from_pdf(uploaded_file):
    customer_pages_text = []

    with pdfplumber.open(uploaded_file) as pdf:
        for page in pdf.pages:
            text = page.extract_text(x_tolerance=1, y_tolerance=3) or ""

            # Group Report pages are summaries; Book5 does not use them.
            if "Commission Group Report" in text:
                continue

            customer_pages_text.append(text)

    full_text = "\n".join(customer_pages_text)
    pay_month, pay_year, ship_month, ship_year = parse_report_period(full_text)

    lines = [clean(line) for line in full_text.splitlines() if clean(line)]

    rows = []
    current_group_name = ""
    current_city = ""
    current_state = ""

    for index, line in enumerate(lines):
        if any(word in line for word in SKIP_WORDS):
            continue

        if line.startswith("Manual Accruals"):
            continue

        parsed_invoice = parse_invoice_line(line)

        if not parsed_invoice:
            continue

        if parsed_invoice["group_name_on_line"]:
            current_group_name = parsed_invoice["group_name_on_line"]

        next_line = lines[index + 1] if index + 1 < len(lines) else ""
        city, state, order_number = parse_city_state_order_line(next_line)

        # Prime PDF prints City/State once at the start of many customer blocks.
        # Carry it forward for later rows in the same block.
        if city:
            current_city = city
        if state:
            current_state = state

        cust_acc_nbr = find_ship_to_account(lines, index + 1)
        cust_name = normalize_customer_name(current_group_name)

        row = {header: "" for header in HEADERS}

        row.update({
            "Supplier_name": "Prime",
            "CustAccNbr": int(cust_acc_nbr) if cust_acc_nbr else "",
            "CustName": cust_name,
            "City": current_city,
            "State": current_state,
            "InvoiceNumber": int(parsed_invoice["invoice_number"]),
            "Qty": 1,
            "UnitCost": parsed_invoice["commission_basis"],
            "Commissions": parsed_invoice["commission_total"],
            "PO_Number": int(order_number) if order_number else "",
            "Pay_Month": pay_month,
            "Pay_Year": pay_year,
            "Ship_Month": ship_month,
            "Ship_Year": ship_year,
        })

        # Force unwanted columns to blank even if they are in the header.
        row["meta_data_json"] = ""
        row["SO_Number"] = ""
        row["Commission_Rate"] = ""
        row["currency"] = ""
        row["rawcustname"] = ""

        rows.append(row)

    # Remove duplicate invoice rows only if PDF extraction repeats the same row.
    seen = set()
    unique_rows = []

    for row in rows:
        key = (
            row.get("InvoiceNumber"),
            row.get("CustAccNbr"),
            row.get("PO_Number"),
            row.get("Commissions"),
        )

        if key not in seen:
            seen.add(key)
            unique_rows.append(row)

    return unique_rows


def create_excel(rows):
    output = BytesIO()
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Extracted Data"

    sheet.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="FFFF00")
    header_font = Font(bold=True)

    for cell in sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for row in rows:
        sheet.append([row.get(header, "") for header in HEADERS])

    for col_idx, header in enumerate(HEADERS, start=1):
        width = min(max(len(header) + 2, 12), 28)
        sheet.column_dimensions[get_column_letter(col_idx)].width = width

    sheet.freeze_panes = "A2"

    workbook.save(output)
    output.seek(0)
    return output


st.title("📊 Prime Conduit PDF Extractor")
st.write("Book5-mapped version. Extracts only customer invoice rows and keeps unwanted columns blank.")

uploaded_files = st.file_uploader(
    "Upload Prime Conduit PDF file(s)",
    type=["pdf"],
    accept_multiple_files=True
)

if uploaded_files:
    if st.button("Extract Data", type="primary"):
        all_rows = []

        with st.spinner("Extracting data..."):
            for uploaded_file in uploaded_files:
                try:
                    file_rows = extract_rows_from_pdf(uploaded_file)
                    all_rows.extend(file_rows)
                    st.success(f"✅ {uploaded_file.name}: {len(file_rows)} rows extracted")
                except Exception as error:
                    st.error(f"❌ {uploaded_file.name}: {error}")

        if all_rows:
            st.dataframe(all_rows, use_container_width=True)

            excel_file = create_excel(all_rows)

            st.download_button(
                label="⬇️ Download Excel",
                data=excel_file,
                file_name="prime_conduit_book5_mapped_output.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
        else:
            st.warning("No invoice rows found. Please make sure the PDF text is selectable.")
else:
    st.info("Upload PDF file(s) to begin.")
